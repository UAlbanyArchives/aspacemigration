import os
from archives_tools import aspace as AS
from archives_tools import uaLocations
from openpyxl import Workbook, load_workbook
import sys

loginData = AS.getLogin()
session = AS.getSession(loginData)
repo = "2"

exportDir = "\\\\romeo\\Collect\\spe\\Tools\\locationUpdates"

count = 0
box = "start"
start = 1
end = 50
while (box != None):
	range = str(start) + "-" + str(end)
	print range
	for box in AS.getContainers(session, repo, range, loginData):
		print (box)
		count = count + 1

		#box = AS.getContainer(session, "/repositories/2/top_containers/30375", loginData)
		if len(box.container_locations) == 1:
			pass
		else:
			if len(box.collection) == 0:
				print "Deleting " + box.long_display_string
				#AS.deleteObject(session, box, loginData)
			elif len(box.collection) > 1:
				AS.pp(box)
				sys.exit()	
			else:
				colID = box.collection[0].identifier
			
				outputFile = os.path.join(exportDir, colID + ".xlsx")
				
				boxLocation = ""
				locCount = 0
				for loc in box.container_locations:
					locCount = locCount + 1
					shelf = AS.getLocation(session, loc.ref, loginData)
					uaLoc = uaLocations.ASpace2Location(shelf.title)
					if locCount > 1:
						boxLocation = boxLocation + "; " + uaLoc
					else:
						boxLocation = boxLocation + uaLoc
				
				if not os.path.isfile(outputFile):
					print ("Writing " + box.type + " " + box.indicator + " to " + colID + ".xlsx")
					wb = Workbook()
					ws = wb.active
					ws["A1"] = "Box URI"
					ws["B1"] = "Box Type"
					ws["C1"] = "Box Number"
					ws["D1"] = "Location"
					ws["A2"] = box.uri
					ws["B2"] = box.type
					ws["C2"] = box.indicator
					ws["D2"] = boxLocation
					
					wb.save(filename = outputFile)
					
				else:
					wb = load_workbook(filename = outputFile)
					ws = wb.active
					
					checkBox = False
					for row in ws.rows:
						if str(row[0].value).strip() == box.uri:
							checkBox = True
					if checkBox == False:
						print ("Writing " + box.type + " " + box.indicator + " to " + colID + ".xlsx")
						newRow = ws.max_row + 1
						ws["A" + str(newRow)] = box.uri
						ws["B" + str(newRow)] = box.type
						ws["C" + str(newRow)] = box.indicator
						ws["D" + str(newRow)] = boxLocation
						
						wb.save(filename = outputFile)
				
	start = end + 1
	end = start + 49
	#print box
	print count